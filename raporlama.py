# importing required modules
import PyPDF2
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
from matplotlib.backends.backend_pdf import PdfPages
# creating a pdf file object
pdfFileObj = open('Tahlil.pdf', 'rb')

# creating a pdf reader object
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

# clear file content
open('results.txt', 'w').close()
# printing number of pages in pdf file
# print(pdfReader.numPages)
for i in range(pdfReader.numPages):

    # creating a page object
    pageObj = pdfReader.getPage(i)

# extracting text from page
    extracted = pageObj.extractText()
    with open('results.txt', 'a') as f:
        f.write(extracted)
    f.close()
# closing the pdf file object
pdfFileObj.close()

# first get all lines from file


degerler = ["Num_Kabul_Zamanı:", "Glukoz(Serum/Plazma)", "eGFR*", "APTT", "PT(INR)", "PTZ(Sn)", "PT(%)", "Kreatinin", "AST", "ALT",
            "Total_Bİlirubin", "Direkt_Bilirubin", "İndirekt_Bilirubin",
            "LDH", "Albumin", "Sodyum", "Potasyum", "Magnezyum", "Klor", "Kalsiyum", "Fosfor", "Ferritin", "Fibrinojen", "D-Dimer",
            "CRP", "Düzeltilmiş_Kalsiyum", "Kalsiyum_X_Fosfor",
            "WBC", "RBC", "HGB", "HCT", "PLT", "MCV", "MCH", "MCHC", "RDW", "NEU#", "LYM#", "EO#", "MON#", "BASO#", "NEU%", "LYM%", "EO%", "MONO%", "BASO%", "MPV", "PCT",
            "PDW", "sO2(ARTERİYEL)", "pH(ARTERİYEL)", "pCO2(ARTERİYEL)", "pO2(ARTERİYEL)", "Hctc(ARTERİYEL)", "ctHb(ARTERİYEL)", "FCOHb(ARTERİYEL)", "FHHb(ARTERİYEL)",
            "FMetHb(ARTERİYEL)", "FO2Hb(ARTERİYEL)", "cK+(ARTERİYEL)", "cNa+(ARTERİYEL)", "cCa2+(ARTERİYEL)", "cCl-(ARTERİYEL)", "cGlu(ARTERİYEL)", "cLac(ARTERİYEL)", "cBase(B)c(ARTERİYEL)",
            "cBase(Ecf)c(ARTERİYEL)", "cHCO3-(P,st)c(ARTERİYEL)", "cHCO3-(P)c(ARTERİYEL)", "nBili(ARTERİYEL)"]
# Python program to read
# file word by word
result_values = []
result_codes = []
# opening the text file
with open('results.txt', 'r') as file2:
    find = False
    # reading each line
    for line in file2:

        line = line.replace("Num.Kabul Zamanı :", ("Num_Kabul_Zamanı:"))
        line = line.replace("( ", ("("))
        line = line.replace(" )", (")"))
        line = line.replace("PT ", ("PT"))
        line = line.replace("PTZ ", ("PTZ"))
        line = line.replace("Glukoz (Serum/Plazma)", ("Glukoz(Serum/Plazma)"))
        line = line.replace("Total Bilirubin", ("Total_Bİlirubin"))
        line = line.replace("Direkt Bilirubin", ("Direkt_Bilirubin"))
        line = line.replace("İndirekt Bilirubin", ("İndirekt_Bilirubin"))
        line = line.replace("Düzeltilmiş Kalsiyum", ("Düzeltilmiş_Kalsiyum"))
        line = line.replace("Kalsiyum X Fosfor", ("Kalsiyum_X_Fosfor"))
        line = line.replace(" (ARTERİYEL)", ("(ARTERİYEL)"))
        line = line.replace("2023 ", ("2023_"))

        for word in line.split():

            if word in degerler:
                find = True
                result_codes.append(str(word))
            elif (find == True):
                word = word.replace("D", "")
                word = word.replace("<", "")
                word = word.replace("(İPTAL)", "")
                word = word.replace("L", "")
                word = word.replace("****", "")
                result_values.append(str(word))
                find = False


df = pd.DataFrame(index=["Glukoz(Serum/Plazma)", "eGFR*", "APTT", "PT(INR)", "PTZ(Sn)", "PT(%)", "Kreatinin", "AST", "ALT",
                         "Total_Bİlirubin", "Direkt_Bilirubin", "İndirekt_Bilirubin",
                         "LDH", "Albumin", "Sodyum", "Potasyum", "Magnezyum", "Klor", "Kalsiyum", "Fosfor", "Ferritin", "Fibrinojen", "D-Dimer",
                         "CRP", "Düzeltilmiş_Kalsiyum", "Kalsiyum_X_Fosfor",
                         "WBC", "RBC", "HGB", "HCT", "PLT", "MCV", "MCH", "MCHC", "RDW", "NEU#", "LYM#", "EO#", "MON#", "BASO#", "NEU%", "LYM%", "EO%", "MONO%", "BASO%", "MPV", "PCT",
                         "PDW", "sO2(ARTERİYEL)", "pH(ARTERİYEL)", "pCO2(ARTERİYEL)", "pO2(ARTERİYEL)", "Hctc(ARTERİYEL)", "ctHb(ARTERİYEL)", "FCOHb(ARTERİYEL)", "FHHb(ARTERİYEL)",
                         "FMetHb(ARTERİYEL)", "FO2Hb(ARTERİYEL)", "cK+(ARTERİYEL)", "cNa+(ARTERİYEL)", "cCa2+(ARTERİYEL)", "cCl-(ARTERİYEL)", "cGlu(ARTERİYEL)", "cLac(ARTERİYEL)", "cBase(B)c(ARTERİYEL)",
                         "cBase(Ecf)c(ARTERİYEL)", "cHCO3-(P,st)c(ARTERİYEL)", "cHCO3-(P)c(ARTERİYEL)", "nBili(ARTERİYEL)"], columns=[])
referanslar = [[74,109],[60,-1],[24,36],[0.8,1.2],[12,16.5],[70,120],[0.70,1.2],
               [-1,40],[-1,41],[-1,1.2],[-1,0.3],[-1,0.8],[135,225],[35,52],[136,145],
               [3.5,5.1],[1.6,2.4],[98,107],[8.6,10.0],[2.5,4.5],[30,400],[2.0,4.0],[-1,550],
               [-1,5],[-1,-1],[-1,-1],[3.98,10.2],[4.69,6.13],[14.1,18.1],[43.5,53.7],[142,424],
               [80,97],[27,31.2],[31.8,35.4],[10,20],[1.78,5.38],[1.32,3.57],[0.04,0.56],[0.12,1.2],
               [0.01,0.08],[34,67.9],[21.8,53.1],[0.8,7],[5.3,12.2],[0.1,1.2],[6.8,10.8],[0.15,0.7],
               [9,19],[-1,-1],[-1,-1],[32.0,48.0],[83,108],[-1,-1],[12,17],[0.5,1.5],[-1,-1],[0.0,1.5],
               [-1,-1],[3.4,4.5],[136,146],[1.15,1.29],[98,106],[-1,-1],[0.5,1.6],[-1,-1],[-1,-1],[-1,-1],
               [-1,-1],[-1,-1],[-1,-1],[-1,-1]]

for i in range(len(result_values)-1):
    if (result_codes[i] == "Num_Kabul_Zamanı:"):
        result_values[i] = result_values[i].replace("_", " ")
        df[result_values[i]] = pd.Series([])
        continue


currentdate = ""
for a in range(len(result_codes)-1):
    if (result_codes[a] == "Num_Kabul_Zamanı:"):
        currentdate = result_values[a]
        continue
    if (result_codes[a] != "Num_Kabul_Zamanı:"):
        df.at[result_codes[a], currentdate] = result_values[a]
        continue
    else:
        continue

cols = list(df.columns)
for col in cols:
    df[[col]] = df[[col]].apply(pd.to_numeric)
df.columns = pd.to_datetime(df.columns, format="%d/%m/%Y %H:%M")
df.sort_index(inplace=True, axis=1)
df.dropna(how="all", axis=1, inplace=True)
df.to_excel("Tahlil.xlsx", na_rep=np.NaN)
var = openpyxl.load_workbook("Tahlil.xlsx")
sh = var.active
print(len(referanslar))

for a in range(2,sh.max_row+1):
    

    madde =""
    veriler = []
    incigun = []
    maddeyeozeltarih = []
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=a, column=j)
        if (j == 1):

            madde = str(cell_obj.value)
        else:
            if (cell_obj.value is None):
                continue
            else:
                veriler.append(cell_obj.value)
                celltarih = sh.cell(row=1, column=j)
                print(celltarih.value.day)
                maddeyeozeltarih.append(str(celltarih.value))
    print(madde)
    
    x_axis = maddeyeozeltarih
    y_axis = veriler
    plt.figure(figsize=(12,6))
    plt.title(madde)
    plt.plot(x_axis, y_axis)
    plt.xlabel("Değerler")
    plt.ylabel("Tarihler")
    plt.xticks(rotation=90, ha='right')
    plt.tight_layout()
    if( referanslar[a-2][1] > 0):
        plt.axhline(y = referanslar[a-2][1], color = 'b', linestyle = '--', label= str(referanslar[a-2][1]))
    if( referanslar[a-2][0] > 0):
        plt.axhline(y = referanslar[a-2][0], color = 'r', linestyle = '--', label= str(referanslar[a-2][0]))
    plt.legend(fontsize = 10)
    for index in range(len(x_axis)):
        plt.text(x_axis[index], y_axis[index], x_axis[index], size=8)
fignums= plt.get_fignums()
p = PdfPages("tum_sonuc_grafigi.pdf")
print(fignums)
figs = [plt.figure(n) for n in fignums]
for fig in figs: 
    fig.savefig(p, format='pdf') 
p.close()  